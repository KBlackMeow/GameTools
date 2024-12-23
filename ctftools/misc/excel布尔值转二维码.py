from openpyxl import load_workbook
import numpy as np
from PIL import Image
from pyzbar import pyzbar
import tqdm

def vist_xlsx(path):
    # 加载Excel文件
    workbook = load_workbook(path)
    sheet = workbook.active  # 或者你选择的工作表索引或名称
    flag = []
    # 便利每一行 每一列
    for row in tqdm.tqdm(sheet.iter_rows()):
        tmp = []
        for cell in row:
            value = cell.value
            font_bold = cell.font.bold  # 检查单元格是否加粗
            if font_bold:
                tmp.append(255)
            else :
                tmp.append(0)
        flag.append(tmp)
    return flag



def read_qrmsg(flag):
    flag = np.array(flag,dtype='uint8')
    # Image.fromarray(flag).save('flag.png')
    # 解码图片中的二维码或条形码
    image = Image.fromarray(flag)
    decoded_objects = pyzbar.decode(image)

    # 输出解码结果
    for obj in decoded_objects:
        print(f'Type: {obj.type}')
        print(f'Data: {obj.data.decode("utf-8")}')
    return obj.data.decode("utf-8")

if __name__ == '__main__':
    flag = vist_xlsx("flag.xlsx")
    flag = read_qrmsg(flag)